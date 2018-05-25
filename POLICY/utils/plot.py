#!/usr/bin/env python
# -*- coding:utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, numbers
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils import get_column_letter
import numpy as np

def tuple_to_coordinate(row, column):
    col_letter = get_column_letter(column)    
    return '{}{}'.format(col_letter, row)

def set_range_style(ws, cell_range, border=Border(), fill=None, font=None, alignment=None, number_format=None, merged=True):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    :param alignment: An openpyxl Alignment object
    :param number_format: An openpyxl Number_format object
    :param merged: Whether merge cells

    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    # default border
    inline = Side(border_style="thin", color="BFBFBF")
    in_border = Border(right=inline , bottom=inline, left=inline, top=inline)
    
    if merged:
        ws.merge_cells(cell_range)
        
    rows = ws[cell_range]

    for row in rows:       
        for c in row:
            c.border = in_border
            if fill:
                c.fill = fill
            if font:
                c.font = font
            if alignment:
                c.alignment = alignment
            if number_format:
                c.number_format = number_format
    
    for c in rows[0]:        
        new_border = Border(left=c.border.left, top=top.top, right=c.border.right, bottom=c.border.bottom)
        c.border = new_border
    for c in rows[-1]:
        new_border = Border(left=c.border.left, top=c.border.top, right=c.border.right, bottom=bottom.bottom)
        c.border = new_border
    for row in rows:
        l = row[0]
        r = row[-1]
        new_border = Border(left=left.left, top=l.border.top, right=l.border.right, bottom=l.border.bottom)
        l.border = new_border
        new_border = Border(left=r.border.left, top=r.border.top, right=right.right, bottom=r.border.bottom)
        r.border = new_border


def set_header(ws, cell_range, name=None, merged=False):
    if name:
        first_cell = ws[cell_range.split(":")[0]]
        first_cell.value = name
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="BCD6EE")
    al = Alignment(horizontal="center", vertical="center")
    font = Font(name=u'微软雅黑', size=11)
    set_range_style(ws, cell_range, border=border, fill=fill, font=font, alignment=al, merged=merged)

def set_title(ws, cell, title):
    ws[cell].font = Font(name=u'微软雅黑', bold=True, size=11)
    ws[cell].value = title

def set_body(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    al = Alignment(horizontal="center", vertical="center")
    font = Font(name=u'微软雅黑', size=11)
    set_range_style(ws, cell_range, border=border, font=font, alignment=al, merged=False)

def set_body_num(ws, cell_range):
    number_format = 'General'
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    al = Alignment(horizontal="center", vertical="center")
    font = Font(name=u'微软雅黑', size=11)
    set_range_style(ws, cell_range, border=border, font=font, alignment=al, number_format=number_format, merged=False)

def plot_table(ws, df, start_cell, title, index_name=u'时间'):
    # plot table title
    set_title(ws, start_cell, title)
    
    # plot table header
    start_row, start_column = coordinate_to_tuple(start_cell)
    
    if (len(df.index.names) == 1):
       # plot columns
       i = 0
       for column_name in df.columns:
           column_cell = ws[tuple_to_coordinate(start_row, start_column+i)]
           cell_range = '%s:%s' % (column_cell.coordinate, column_cell.coordinate)
           set_header(ws, cell_range, name=column_name, merged=False)
           i = i + 1
    elif (len(df.index.names) == 2):
        # plot index column
        index_cell = ws[tuple_to_coordinate(start_row+1, start_column)]
        cell_range = '%s:%s' % (index_cell.coordinate, index_cell.coordinate)
        set_header(ws, cell_range, name=index_name, merged=False)
        # plot columns
        i = 1
        for column_name in df.index.levels[0]:
            column_cell = ws[tuple_to_coordinate(start_row+1, start_column+i)]
            cell_range = '%s:%s' % (column_cell.coordinate, column_cell.coordinate)
            set_header(ws, cell_range, name=column_name, merged=False)
            i = i + 1
    else:
        # plot index column
        index_cell_lt = ws[tuple_to_coordinate(start_row + 1, start_column)]
        index_cell_rb = ws[tuple_to_coordinate(start_row + 2, start_column)]
        cell_range = '%s:%s' % (index_cell_lt.coordinate, index_cell_rb.coordinate)
        set_header(ws, cell_range, name=index_name, merged=True)
        
        #plot top level column
        second_column_size = len(df.index.levels[1])
        top_column_row = start_row + 1
        top_column_column = start_column + 1
        second_column_row = start_row + 2
        for top_column_name in df.index.levels[0]:
            column_start_cell = ws[tuple_to_coordinate(top_column_row, top_column_column)]
            column_end_cell = ws[tuple_to_coordinate(top_column_row, top_column_column + second_column_size - 1)]
            top_column_cell_range = '%s:%s' % (column_start_cell.coordinate, column_end_cell.coordinate)
            set_header(ws, top_column_cell_range, name=top_column_name, merged=True)
            
            second_column_start_cell = ws[tuple_to_coordinate(second_column_row, top_column_column)]
            second_column_end_cell = ws[tuple_to_coordinate(second_column_row, top_column_column + second_column_size - 1)]
            second_column_cell_range = '%s:%s' % (second_column_start_cell.coordinate, second_column_end_cell.coordinate)
            # plot second columns
            second_column_column = top_column_column
            for second_column_name in df.index.levels[1]:
                second_column_cell = ws[tuple_to_coordinate(second_column_row, second_column_column)]
                second_column_cell.value = second_column_name
                second_column_column = second_column_column + 1
            # plot border
            set_header(ws, second_column_cell_range, merged=False)

            top_column_column = top_column_column + second_column_size
    
    # plot body
    if (len(df.index.names) == 1):
        # set body style
        value_row = start_row + 1
        value_column = start_column
       
        for i in range(0, df.shape[0]):
            s = df.loc[i]
            j = 0
            for d in df.columns:
                v = s.get(d)
                value_cell = ws[tuple_to_coordinate(value_row+i, value_column+j)]
                value_cell_range = '%s:%s' % (value_cell.coordinate, value_cell.coordinate)
                set_body(ws, value_cell_range)
                value_cell.value = v
                if isinstance(value_cell.value, np.float64):
                    value_cell.number_format = numbers.FORMAT_NUMBER
                j = j + 1
    elif (len(df.index.names) == 2):
        write_date = True
        index_row = start_row + 2
        index_column = start_column
        value_column = start_column + 1
        
        # set date column style
        index_start_cell = ws[tuple_to_coordinate(index_row, index_column)]
        index_end_cell = ws[tuple_to_coordinate(index_row + len(df.index.levels[1]) - 1, index_column)]
        index_cell_range = '%s:%s' % (index_start_cell.coordinate, index_end_cell.coordinate)
        set_body(ws, index_cell_range)
        
        for c0 in df.index.levels[0]:
            s0 = df.loc[c0]

            # set body style
            value_row = start_row + 2
            value_start_cell = ws[tuple_to_coordinate(value_row, value_column)]
            value_end_cell = ws[tuple_to_coordinate(value_row + len(df.index.levels[1]) - 1, value_column)]
            value_cell_range = '%s:%s' % (value_start_cell.coordinate, value_end_cell.coordinate)
            set_body_num(ws, value_cell_range)
            

            for c1 in df.index.levels[1]:
                v = s0.loc[c1][0] if c1 in s0.index else ''
                if write_date:
                    index_cell = ws[tuple_to_coordinate(index_row, index_column)]
                    index_cell.value = c1
                    index_row = index_row + 1
                value_cell = ws[tuple_to_coordinate(value_row, value_column)]
                value_cell.value = v
                if isinstance(value_cell.value, np.float64):
                    if value_cell.value < 1:
                        value_cell.number_format =  numbers.FORMAT_PERCENTAGE_00
                    else:
                        value_cell.number_format = numbers.FORMAT_NUMBER_00
                        try:
                            if int(value_cell.value) == value_cell.value:
                                value_cell.number_format = numbers.FORMAT_NUMBER
                        except:
                            pass
                value_row = value_row + 1
            
            write_date = False
            value_column = value_column + 1
    else:
        
        write_index = True
        index_row = start_row + 3
        index_column = start_column
        value_column = start_column + 1
        
        # set date column style
        index_start_cell = ws[tuple_to_coordinate(index_row, index_column)]
        index_end_cell = ws[tuple_to_coordinate(index_row + len(df.index.levels[2]) - 1, index_column)]
        index_cell_range = '%s:%s' % (index_start_cell.coordinate, index_end_cell.coordinate)
        set_body(ws, index_cell_range)
        
        for c0 in df.index.levels[0]:
            s0 = df.loc[c0]
            
            value_row = start_row + 3
            # set body styel
            value_start_cell = ws[tuple_to_coordinate(value_row, value_column)]
            value_end_cell = ws[tuple_to_coordinate(value_row + len(df.index.levels[2]) - 1, value_column + len(df.index.levels[1]) - 1)]
            value_cell_range = '%s:%s' % (value_start_cell.coordinate, value_end_cell.coordinate)
            set_body_num(ws, value_cell_range)
            
            for c1 in df.index.levels[1]:
                s1 = s0.loc[c1]
                
                value_row = start_row + 3
                for c2 in df.index.levels[2]:                  
                    if write_index:
                        index_cell = ws[tuple_to_coordinate(index_row, index_column)]
                        index_cell.value = c2
                        index_row = index_row + 1
                    
                    v = s1.loc[c2][0] if c2 in s1.index else ''
                    value_cell = ws[tuple_to_coordinate(value_row, value_column)]
                    value_cell.value = v
                    if isinstance(value_cell.value, np.float64):
                        if value_cell.value < 1:
                            value_cell.number_format =  numbers.FORMAT_PERCENTAGE_00 
                        else:
                            value_cell.number_format = numbers.FORMAT_NUMBER_00
                            try:
                                if int(value_cell.value) == value_cell.value:
                                    value_cell.number_format = numbers.FORMAT_NUMBER
                            except:
                                pass
                    value_row = value_row + 1
                
                write_index = False
                value_column = value_column + 1


